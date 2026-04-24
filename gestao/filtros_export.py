# ==========================================
# gestao/filtros_export.py
#
# Dois utilitários globais:
#   1. get_filtros_base(request) — extrai e valida os parâmetros
#      de filtro padrão presentes em qualquer listagem.
#   2. exportar_pdf / exportar_excel — geram HttpResponse com o
#      arquivo para download a partir de qualquer lista de dicts.
# ==========================================

from django.http import HttpResponse
from django.utils.dateparse import parse_date
from datetime import date, timedelta


# ══════════════════════════════════════════════════════════════════════════════
# 1. HELPER DE FILTROS
# ══════════════════════════════════════════════════════════════════════════════

def get_filtros_base(request):
    """
    Extrai e normaliza os parâmetros de filtro padrão da query string.

    Parâmetros aceitos:
      ?data_inicio=YYYY-MM-DD   — início do período
      ?data_fim=YYYY-MM-DD      — fim do período
      ?atalho=hoje|semana|mes|mes_anterior|ano  — atalhos de período
      ?cliente_id=INT
      ?fornecedor_id=INT
      ?vendedor_id=INT
      ?produto_id=INT
      ?status=STRING

    Retorna um dict com todos os valores já convertidos para os
    tipos corretos (date, int, str). Valores ausentes ficam como None.
    """
    params = request.query_params

    # ── Atalhos de período ────────────────────────────────────────────────
    atalho = params.get('atalho')
    hoje = date.today()
    data_inicio = None
    data_fim = None

    if atalho == 'hoje':
        data_inicio = hoje
        data_fim = hoje
    elif atalho == 'semana':
        data_inicio = hoje - timedelta(days=hoje.weekday())
        data_fim = hoje
    elif atalho == 'mes':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    elif atalho == 'mes_anterior':
        primeiro_do_mes = hoje.replace(day=1)
        ultimo_mes = primeiro_do_mes - timedelta(days=1)
        data_inicio = ultimo_mes.replace(day=1)
        data_fim = ultimo_mes
    elif atalho == 'ano':
        data_inicio = hoje.replace(month=1, day=1)
        data_fim = hoje
    else:
        # Datas explícitas têm prioridade sobre atalhos
        raw_ini = params.get('data_inicio')
        raw_fim = params.get('data_fim')
        data_inicio = parse_date(raw_ini) if raw_ini else None
        data_fim = parse_date(raw_fim) if raw_fim else None

    # ── IDs de referência ─────────────────────────────────────────────────
    def _int(key):
        v = params.get(key)
        try:
            return int(v) if v else None
        except (ValueError, TypeError):
            return None

    return {
        'data_inicio': data_inicio,
        'data_fim':    data_fim,
        'cliente_id':  _int('cliente_id'),
        'fornecedor_id': _int('fornecedor_id'),
        'vendedor_id': _int('vendedor_id'),
        'produto_id':  _int('produto_id'),
        'status':      params.get('status') or None,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 2. EXPORTAÇÃO PDF
# ══════════════════════════════════════════════════════════════════════════════

def exportar_pdf(titulo, colunas, linhas, filename='relatorio.pdf'):
    """
    Gera um PDF simples com cabeçalho e tabela.

    Args:
        titulo (str): Título exibido no topo do documento.
        colunas (list[str]): Lista com os nomes das colunas.
        linhas (list[list]): Dados — cada elemento é uma lista de valores
                             na mesma ordem das colunas.
        filename (str): Nome do arquivo enviado ao navegador.

    Returns:
        HttpResponse com Content-Type application/pdf.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io

    buffer = io.BytesIO()
    largura, altura = landscape(A4) if len(colunas) > 6 else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(largura, altura),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'Titulo', parent=styles['Heading1'],
        fontSize=14, textColor=colors.HexColor('#1A6B3C'),
        alignment=TA_CENTER, spaceAfter=12,
    )
    estilo_rodape = ParagraphStyle(
        'Rodape', parent=styles['Normal'],
        fontSize=8, textColor=colors.grey, alignment=TA_CENTER,
    )

    from datetime import datetime
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')

    elementos = [
        Paragraph(titulo, estilo_titulo),
        Paragraph(f'Gerado em {agora}', estilo_rodape),
        Spacer(1, 0.4 * cm),
    ]

    # Monta dados da tabela: cabeçalho + linhas
    dados_tabela = [colunas] + [[str(v) if v is not None else '' for v in linha] for linha in linhas]

    n_cols = len(colunas)
    col_width = (largura - 3 * cm) / n_cols

    tabela = Table(dados_tabela, colWidths=[col_width] * n_cols, repeatRows=1)
    tabela.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A6B3C')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        # Linhas de dados
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ALIGN',      (0, 1), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabela)

    if not linhas:
        elementos.append(Spacer(1, 0.5 * cm))
        elementos.append(Paragraph('Nenhum registro encontrado para os filtros selecionados.', styles['Normal']))

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# 3. EXPORTAÇÃO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def exportar_excel(titulo, colunas, linhas, filename='relatorio.xlsx'):
    """
    Gera um arquivo Excel (.xlsx) com cabeçalho formatado e dados.

    Args:
        titulo (str): Usado como nome da aba e título da planilha.
        colunas (list[str]): Nomes das colunas.
        linhas (list[list]): Dados na mesma ordem das colunas.
        filename (str): Nome do arquivo enviado ao navegador.

    Returns:
        HttpResponse com Content-Type application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    # Nome da aba: max 31 chars (limite do Excel)
    ws.title = titulo[:31]

    # ── Estilos ────────────────────────────────────────────────────────────
    verde = '1A6B3C'
    cinza_claro = 'F5F5F5'

    fonte_cabecalho = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    fill_cabecalho  = PatternFill('solid', fgColor=verde)
    alinha_centro   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alinha_esq      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    borda_fina      = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    fill_par   = PatternFill('solid', fgColor=cinza_claro)
    fill_impar = PatternFill('solid', fgColor='FFFFFF')

    # ── Título ─────────────────────────────────────────────────────────────
    from datetime import datetime
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    cell_titulo = ws.cell(row=1, column=1, value=titulo)
    cell_titulo.font = Font(name='Arial', bold=True, size=13, color=verde)
    cell_titulo.alignment = alinha_centro
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(colunas))
    cell_data = ws.cell(row=2, column=1, value=f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    cell_data.font = Font(name='Arial', size=8, color='888888')
    cell_data.alignment = alinha_centro
    ws.row_dimensions[2].height = 14

    # Linha em branco separadora
    ws.row_dimensions[3].height = 6

    # ── Cabeçalho das colunas ──────────────────────────────────────────────
    LINHA_CABECALHO = 4
    for col_idx, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=LINHA_CABECALHO, column=col_idx, value=nome)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinha_centro
        cell.border = borda_fina
    ws.row_dimensions[LINHA_CABECALHO].height = 20

    # ── Dados ──────────────────────────────────────────────────────────────
    for row_idx, linha in enumerate(linhas, start=LINHA_CABECALHO + 1):
        fill = fill_par if (row_idx % 2 == 0) else fill_impar
        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font  = Font(name='Arial', size=9)
            cell.fill  = fill
            cell.alignment = alinha_esq
            cell.border = borda_fina
        ws.row_dimensions[row_idx].height = 16

    # ── Ajusta largura das colunas ─────────────────────────────────────────
    for col_idx, nome in enumerate(colunas, start=1):
        letra = get_column_letter(col_idx)
        # Largura baseada no maior conteúdo (colunas + dados), limitada a 40
        max_len = len(str(nome))
        for linha in linhas:
            if col_idx - 1 < len(linha):
                max_len = max(max_len, len(str(linha[col_idx - 1] or '')))
        ws.column_dimensions[letra].width = min(max_len + 4, 40)

    # ── Freeze pane no cabeçalho ───────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=LINHA_CABECALHO + 1, column=1)

    # ── Salva e retorna ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(buffer, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# 4. HELPER: formata dados de relatório para exportação
# ══════════════════════════════════════════════════════════════════════════════

def preparar_exportacao(dados, mapeamento):
    """
    Converte uma lista de dicts em (colunas, linhas) prontos para
    exportar_pdf() ou exportar_excel().

    Args:
        dados (list[dict]): Resultado de qualquer função de relatório.
        mapeamento (list[tuple]): Lista de (chave_do_dict, 'Nome da Coluna').

    Exemplo:
        mapeamento = [
            ('cliente', 'Cliente'),
            ('valor',   'Valor R$'),
            ('vencimento', 'Vencimento'),
        ]
    """
    colunas = [col for _, col in mapeamento]
    linhas  = []
    for item in dados:
        linha = [item.get(chave, '') for chave, _ in mapeamento]
        linhas.append(linha)
    return colunas, linhas
